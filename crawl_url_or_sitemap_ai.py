#!/usr/bin/env python
"""\
Sitemap crawler using ai: given root url for a sitemap, crawl all web pages in the sitemap and return result in markdown format.
# Source: https://www.reddit.com/r/DataHoarder/comments/1iknxwj/crawl4ai_code_example/, also see https://www.youtube.com/watch?v=JWfNLF_g_V0&t=671s
DONE:
    - Add restartability: save list of crawled pages to a file
        - Add a flag to force crawl of previously crawled pages
TODO:
    - Add the url content to be saved to files with encoded url as filename
    - Add a regular expression of content types to save when crawling (only matching content types will be saved)
TEST:
    - User query to only download web content relevant to the query
"""
__author__ = "Ram P"
__email__ = "rambpol@yahoo.com"
__version__ = "0.1.0"
# __maintainer__ = __author__
# __status__ = "Development"
# __copyright__  = "Copyright 2023, " + __author__

import os
import sys
import argparse
from urllib.parse import urljoin, urlparse, urldefrag, quote as urllib_quote
from tldextract import extract as tld_extract
from openpyxl import Workbook, load_workbook
from pydantic import BaseModel, HttpUrl, ValidationError
import uuid
import asyncio
from crawl4ai import AsyncWebCrawler, BrowserConfig, CrawlResult, CrawlerRunConfig, CacheMode
from crawl4ai.content_filter_strategy import BM25ContentFilter, PruningContentFilter
from crawl4ai.markdown_generation_strategy import DefaultMarkdownGenerator
# from crawl4ai.rate_limit_strategy import RateLimitConfig

import json
from datetime import datetime
import xml.etree.ElementTree as ET
import aiohttp
from aiofiles import open as aio_open
from asyncio import Lock
json_file_lock = Lock()
txt_file_lock = Lock()
error_file_lock = Lock()
wb_lock = Lock()  # lock for writing to excel file

import logging
LOGGER_NAME =  os.path.splitext(os.path.basename(__file__))[0]
logger = logging.getLogger(name=LOGGER_NAME) # root logger by default, pass LOGGER_NAME for script specific log file
logging.raiseExceptions = False

# Configuration
ROOT_URL_DEFAULT = '' # 'https://crawl4ai.com'
SITEMAP_URL_DEFAULT = "https://ai.pydantic.dev/sitemap.xml"  # 'https://docs.crawl4ai.com/sitemap.xml' # "https://www.cnn.com/sitemap.xml"  # your sitemap URL
MAX_DEPTH_DEFAULT = 10  # Limit web pages crawl depth
FORCE_CRAWL = True  # Force crawl of previously crawled pages
BATCH_SIZE = 2  # Number of concurrent crawls
REQUEST_DELAY = 1  # Delay between requests (seconds)
MAX_FILE_SIZE_MB = 20  # Max file size before creating a new one
OUTPUT_DIR_DEFAULT = "output"  # Directory to store multiple output files
EXCEL_FILENAME_SUFFIX = ".xlsx"  # Suffix for excel files
VISITED_URLS_COLUMN = 0
VISITED_URL_DEPTH_COLUMN = 1
VISITED_URL_CONTENT_SIZE_COLUMN = 2
RETRY_LIMIT = 1  # Retry failed URLs once
LOG_FILE = os.path.join(OUTPUT_DIR_DEFAULT, "crawler_log.txt")  # Log file for general logging
ERROR_LOG_FILE = os.path.join(OUTPUT_DIR_DEFAULT, "logfile.txt")  # Log file for failed URLs
BROWSER_TYPE = "chromium",  # or "firefox" or "webkit"
CRAWL4AI_BROWSER_DATA_DIR = "./cral4ai_browser_data"  # to persist cookies, etc.

# Ensure output directory exists
os.makedirs(OUTPUT_DIR_DEFAULT, exist_ok=True)

async def log_message(message, file_path=LOG_FILE):
    """Log messages to a log file and print them to the console."""
    if message.startswith("❌") or message.startswith("✅"): # or message.startswith("🔍"): # or message.startswith("⚠️") or message.startswith("🔄"):
        # prepend message with timestamp
        timestamp = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
        message = f"{timestamp} - {message}" 
    async with aio_open(file_path, "a", encoding="utf-8") as f:
        await f.write(message + "\n")
    print(message)

async def fetch_sitemap(sitemap_url):
    """Fetch and parse sitemap.xml to extract all URLs."""
    try:
        async with aiohttp.ClientSession() as session:
            async with session.get(sitemap_url) as response:
                if response.status == 200:
                    xml_content = await response.text()
                    root = ET.fromstring(xml_content)
                    urls = [elem.text for elem in root.findall(".//{http://www.sitemaps.org/schemas/sitemap/0.9}loc")]

                    if not urls:
                        await log_message("❌ No URLs found in the sitemap.")
                    return urls
                else:
                    await log_message(f"❌ Failed to fetch sitemap: HTTP {response.status}")
                    return []
    except Exception as e:
        await log_message(f"❌ Error fetching sitemap: {str(e)}")
        return []

async def get_file_size(file_path):
    """Returns the file size in MB."""
    if os.path.exists(file_path):
        return os.path.getsize(file_path) / (1024 * 1024)  # Convert bytes to MB
    return 0

async def get_new_file_path(file_prefix, extension):
    """Generates a new file path when the current file exceeds the max size."""
    index = 1
    while True:
        file_path = os.path.join(OUTPUT_DIR_DEFAULT, f"{file_prefix}_{index}.{extension}")
        if not os.path.exists(file_path) or await get_file_size(file_path) < MAX_FILE_SIZE_MB:
            return file_path
        index += 1

async def write_to_json(data, file_prefix, extension):
    """Writes a single JSON object as a line to a file, ensuring size limit."""
    async with json_file_lock:
        file_path = await get_new_file_path(file_prefix, extension)
        timestamp = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
        async with aio_open(file_path, "a", encoding="utf-8") as f:
            await f.write(json.dumps(data, ensure_ascii=False) + "\n")

async def write_to_txt(data, file_prefix):
    """Writes extracted content to a TXT file while managing file size."""
    async with txt_file_lock:
        file_path = await get_new_file_path(file_prefix, "txt")
        timestamp = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
        async with aio_open(file_path, "a", encoding="utf-8") as f:
            await f.write(f"Snapshot at: {timestamp}\nURL: {data['url']}\nTitle: {data['title']}\nContent:\n{data['content']}\n\n{'='*80}\n\n")

async def write_failed_url(url):
    """Logs failed URLs to a separate error log file."""
    async with error_file_lock:
        async with aio_open(ERROR_LOG_FILE, "a", encoding="utf-8") as f:
            await f.write(url + "\n")

def add_chaild_url(url, parent_url, root_domain, depth, queue, visited_urls):
    """Adds a child url to the queue if not already visited and within depth limit."""
    if len(url) == 0 or url == '.':
        return -1, 'Empty URL'
    parent_url_parsed = urlparse(parent_url)
    url_parsed = urlparse(url)
    if url_parsed.path.endswith(':'):  # things like javascript: or mailto:
        return -2, 'javascript or alternative redirect URL'
    if not 'http' in url_parsed.scheme.lower():
        if len(url_parsed.scheme) > 0:
            return -3, 'Invalid URL scheme'
    if len(url_parsed.netloc) == 0:
        if url.startswith('#'):
            return -4, 'Fragment URL'
        else:
            url = parent_url_parsed.scheme + '://' + parent_url_parsed.netloc + '/' + url
            url_parsed = urlparse(url)
            logger.debug(f'-Reset domain: <{url}>: urlparsed: {url_parsed}: in url <{parent_url}>')
    url = urldefrag(url).url
    if root_domain in url_parsed.netloc:
        if (url, depth+1) in queue:
            # logger.debug(f'--Discarding link <{url}>: already in queue')
            return -5, 'Already in queue'
        else:
            # the URL discovered is new
            if url not in visited_urls:
                # logger.info(f'---Appending new link <{url}> at depth {depth+1} from url <{parent_url}>: queue length={len(queue)}')
                queue.append((url, depth+1))
                return 1, 'New URL added to queue'
            else:
                # logger.debug(f'--Discarding link <{url}>: already visited')
                return -6, 'Already visited URL'
    else:
        # url_path_split = os.path.splitext(url_parsed.path)
        # if len(url_path_split[1]) == 0 or not url_path_split[1] in EXTERNAL_FILE_TYPES_TO_DOWNLOAD:
        #     logger.debug(f'--Discarding link <{url}>: external domain')
        # else: 
        #     if url not in visited_urls and url not in urls:
        #         logger.info(f'---Appending new link <{url}> at depth {depth+1} from url <{parent_url}>: queue length={len(queue)}')
        #         urls.add(url)
        #     else:
        #         logger.debug(f'--Discarding link <{url}>: already visited or queued to visit')
        return -7, 'External domain URL'
    return 0, 'Unknown error adding child URL'

async def crawl_url(url, root_domain, crawler, run_config, depth, max_depth, semaphore, visited_urls, queue, 
                    total_urls, retry_count=0, previously_crawled_urls=set(), crawled_urls_filename=None, force_crawl=False):
    if url in visited_urls:
        await log_message(f"⚠️ Skipping {url} as it is already visited.")
        return
    if depth > max_depth:
        await log_message(f"⚠️ Skipping {url} at depth {depth} as max depth {max_depth} reached.")
        return
    visited_urls.add(url)
    """Crawls a single URL, handles retries, logs failed URLs, and extracts child links."""
    async with semaphore:
        await asyncio.sleep(REQUEST_DELAY)  # Rate limiting
        await log_message(f"🔍 {len(visited_urls)}/{(len(queue)+len(visited_urls))} - Crawling {url} at depth {depth}...")
        async with crawler:
            try:
                result = await crawler.arun(url=url, config=run_config)
                if result.success:
                    data = {
                        "url": result.url,
                        "title": result.markdown_v2.raw_markdown.split("\n")[0] if result.markdown_v2.raw_markdown else "No Title",
                        "content": result.markdown_v2.fit_markdown,
                    }

                    if url not in previously_crawled_urls:
                        # Save extracted data
                        # await write_to_json(data, os.path.join(root_domain, root_domain + "_data"), "jsonl")
                        await write_to_txt(data, os.path.join(root_domain, root_domain + "_data")
                        if crawled_urls_filename is not None:
                            async with wb_lock:
                                wb = load_workbook(crawled_urls_filename)
                                ws = wb.active
                                ws.append([url, depth, str(len(data["content"]))]) 
                                logger.debug(f'Saved {url} to {crawled_urls_filename}')
                                wb.save(crawled_urls_filename)
                                wb.close()

                    await log_message(f"✅ {len(visited_urls)}/{(len(queue)+len(visited_urls))} - Successfully crawled: {url}")

                    if depth >= max_depth:
                        await log_message(f"⚠️ Skipping child links for {url} due to max depth reached.")
                    else:
                        # Extract and queue child pages
                        for link in result.links.get("internal", []):
                            href = link["href"]
                            absolute_url = urljoin(url, href)  # Convert to absolute URL
                            ret, message = add_chaild_url(absolute_url, url, root_domain, depth, queue, visited_urls)
                            if ret != 1:
                                await log_message(f"⚠️ Skipping child url {absolute_url} from {url} due to {message}")
                                pass
                else:
                    await log_message(f"⚠️ Failed to extract content from: {url}")

            except Exception as e:
                if retry_count < RETRY_LIMIT:
                    await log_message(f"🔄 Retrying {url} (Attempt {retry_count + 1}/{RETRY_LIMIT}) due to error: {str(e)}")
                    await crawl_url(url, root_domain, crawler, run_config, depth, max_depth, semaphore, visited_urls, queue, total_urls, 
                                    retry_count + 1, previously_crawled_urls, crawled_urls_filename, force_crawl)
                else:
                    await log_message(f"❌ Skipping {url} after {RETRY_LIMIT} failed attempts.")
                    await write_failed_url(url)

async def crawl_urls(urls, root_domain, max_depth=MAX_DEPTH_DEFAULT, batch_size=BATCH_SIZE, user_query=None, force_crawl=False):  
    """Crawls all URLs in a list following child links up to max depth with parallelism of batch_size."""
    if not urls:
        await log_message("❌ No URLs to crawl. Exiting.")
        return

    total_urls = len(urls)  # Total number of URLs to process
    visited_urls = set()
    previously_crawled_urls = set()  # Set to store previously crawled URLs
    queue = [(url, 0) for url in urls]

    crawled_urls_filename = root_domain + EXCEL_FILENAME_SUFFIX
    crawled_urls_filename = urllib_quote(crawled_urls_filename, safe='', encoding=None, errors=None)
    crawled_urls_filename = os.path.join(OUTPUT_DIR_DEFAULT, crawled_urls_filename)

    logger.info(f'crawl_filename={crawled_urls_filename}')
    if os.path.exists(crawled_urls_filename):
        wb = load_workbook(crawled_urls_filename)
        ws = wb.active
        total_urls_crawled = ws.max_row - 1
        for row in ws.iter_rows(min_row=2, values_only=True):
            previously_crawled_urls.add(row[VISITED_URLS_COLUMN])
        logger.info(f'Loaded {total_urls_crawled} previously crawled urls from {crawled_urls_filename}')
        wb.close()
        if not force_crawl:
            visited_urls = previously_crawled_urls.copy()  # make a copy of previously crawled URLs
    else:
        logger.info(f'No previously crawled urls file for {root_domain}, creating new file: {crawled_urls_filename}')
        wb = Workbook()
        # ws = wb.create_sheet()
        ws = wb.active # default sheet
        ws.append(['URL', 'Depth', 'Content Size'])
        wb.save(crawled_urls_filename)
        wb.close()

    semaphore = asyncio.Semaphore(batch_size)  # Concurrency control
    crawlers = []
    run_configs = []
    for index in range(batch_size): # create a list of crawlers
        browser_cfg = BrowserConfig(
                browser_type=BROWSER_TYPE, 
                headless=True,
                viewport_width=1280,
                viewport_height=720,
                # proxy="http://user:pass@myproxy:8080",
                use_persistent_context=True,  # to persist cookies, etc.
                user_data_dir=CRAWL4AI_BROWSER_DATA_DIR,  # to persist cookies, etc.
                ignore_https_errors=True,   # dont care if invalid certificates when crawling web pages
                java_script_enabled=True,  # Set to False to disable javascript
                cookies=[],  # list of cookies to be used in the browser: each a dict like {"name": "session", "value": "...", "url": "..."}
                headers={},  # Set headers to be used in the browser
                text_mode=False, # Set to True to disable images, css, fonts, etc.  (want images)
        )
        crawlers.append(AsyncWebCrawler(browser_config=browser_cfg, 
            # rate_limit=RateLimitConfig(request_delay=REQUEST_DELAY, max_requests_per_second=1),
            retry_limit=RETRY_LIMIT, 
            max_file_size_mb=MAX_FILE_SIZE_MB,
            save_media_files=True,  # Set to True to save media files (images, audio and video) found on webpages
            save_crawl_to_file=True,  # Set to True to save list of crawled pages to a file
            content_types='.*',  # regular expression of content types to save when crawling (only matching content types will be saved)
        ))
        # Configure the run
        run_configs.append(CrawlerRunConfig(
            # cache_mode=CacheMode.ENABLED,
            cache_mode=CacheMode.BYPASS,
            session_id="crawl4ai_session_" + root_domain,  # + str(uuid.uuid1()),  # unique session id based on network address and time of the machine it is running on
            css_selector=None, # "main.article",
            excluded_selector="script, style, nav, footer",
            excluded_tags=["script", "style", "nav", "footer"],
            keep_data_attributes=True, # want to keep data-* attributes
            # wait_for="css:.main-content",  # condition before content extraction: wait for a CSS ("css:selector") like "css:.article-loaded", or JS ("js:() => bool") 
            wait_for_images=True,
            delay_before_return_html=0.1,  # 0 seconds
            check_robots_txt=True, # respect robots.txt rules
            wait_until="networkidle",  # "networkidle0" or "networkidle2" or "load" or "domcontentloaded"
            page_timeout=60000,  # milliseconds
            semaphore_count=1,  # crawl concurrency is controlled external to the crawler
            word_count_threshold=15,
            # screenshot=True,
            # enable_rate_limiting=True,
            # rate_limit_config=RateLimitConfig(
                # base_delay=(1.0, 3.0),
                # max_delay=60.0,
                # max_retries=3,
                # rate_limit_codes=[429, 503],
            # ),
            # memory_threshold_percent=70.0,
            # check_interval=1.0,
            # max_session_permit=20,
            # display_mode="DETAILED",
            # stream=True,
            js_code = None,  # JavaScript to run after load. E.g. "document.querySelector('button')?.click();" or "window.scrollTo(0, document.body.scrollHeight);"
            js_only = False,  # Set to True to run only JavaScript code
            ignore_body_visibility=False,  # Set to True to ignore body visibility
            scan_full_page=False,  # Set to False to scan only the visible part of the page
            scroll_delay=0.5,  # 0.5 seconds between scrolls when full page is not visible
            process_iframes=True,  # Set to False to ignore iframes
            remove_overlay_elements=True,  # Set to False to keep overlay elements
            simulate_user=True,  # Set to False to disable user simulation (to avoid bot detection)
            override_navigator=True,  # Set to False to use the default navigator object (Override navigator properties in JS for stealth.)
            magic=True,  # Set to False to disable magic mode: Automatic handling of popups/consent banners. Experimental.
            image_score_threshold=0.5,  # Image score threshold for image extraction
            exclude_external_images=True,  # Set to False to keep external images
            # exclude_social_media_domains=[],  # A default list can be extended. Any link to these domains is removed from final output.
            exclude_external_links=True,  # Set to False to keep external links
            exclude_social_media_links=True,  # Set to False to keep social media links
            exclude_domains=[],  # A default list can be extended. Any link to these domains is removed from final output.
            extraction_strategy=None,  # "readability" or "goose" or "newspaper" or "lxml" or "html5lib" or "beautifulsoup4"
            # extraction_strategy_config=None,  # {"readability": {"min_text_length": 100}}
            # extraction_strategy_config={"goose": {"use_meta_language": False}},
            # extraction_strategy_config={"newspaper": {"use_meta_language": False}},
            markdown_generator=DefaultMarkdownGenerator(
                content_filter=PruningContentFilter(threshold=0.5, threshold_type="fixed")
            ),
            # markdown_generator="MARKDOWNIT",  # "MARKDOWNIT" or "MISTUNE" or "MARKDOWN"
            # markdown_generator=DefaultMarkdownGenerator(),
            # markdown_generator_config={"mistune": {"escape": False}},
            # markdown_generator_config={"markdown": {"escape": False}},
            # PruningContentFilter(
            #     min_text_length=100,    # Minimum text length in a paragraph
            #     min_word_count=5,       # Minimum word count in a paragraph (after removing stop words)
            #     min_sentence_count=1,   # Minimum sentence count in a paragraph (after removing stop words)
            #     min_paragraph_count=1,  # Minimum paragraph count in the content
            #     min_content_length=1000,  # Minimum content length
            #     min_content_word_count=100,  # Minimum content word count
            #     min_content_sentence_count=5,  # Minimum content sentence count
            #     min_content_paragraph_count=5,  # Minimum content paragraph count
            # )
            content_filter = BM25ContentFilter(
                user_query=user_query,  # User query to adjust BM25 scores
                # Adjust for stricter or looser results
                bm25_threshold=1.2  
            ) if user_query else None,
        ))

    time_start = datetime.now()
    while len(queue)>0:
        tasks = []
        batch = queue[:batch_size]
        queue = queue[batch_size:]
        index = 0
        for url, depth in batch:
            if not url in visited_urls:
                tasks.append(crawl_url(url, root_domain, crawlers[index], run_configs[index], 
                                       depth, max_depth, semaphore, visited_urls, queue, 
                                       total_urls, 0, previously_crawled_urls, crawled_urls_filename, force_crawl))
                index += 1

        await asyncio.gather(*tasks)
    
    await log_message(f"✅ {len(visited_urls)}/{(len(queue)+len(visited_urls))} - Successfully crawled {total_urls} URLs of {root_domain}.")
    time_end = datetime.now()
    elapsed_time = time_end - time_start
    elapsed_hours, remainder = divmod(elapsed_time.total_seconds(), 3600)
    elapsed_minutes, _ = divmod(remainder, 60)
    await log_message(f"✅ Crawling completed in {int(elapsed_hours)} hours and {int(elapsed_minutes)} minutes.")

async def main(args):
    # Clear previous logs
    async with aio_open(LOG_FILE, "w") as f:
        await f.write("")
    async with aio_open(ERROR_LOG_FILE, "w") as f:
        await f.write("")
    
    urls = []
    root_domain = None

    if args.ROOT_URL and len(args.ROOT_URL) > 0:
        if not root_domain:
            root_domain = get_root_domain(args.ROOT_URL)
        else:
            validate(args.ROOT_URL)
        urls.append(args.ROOT_URL)
        await log_message(f"✅ Added root URL {args.ROOT_URL} with root domin {root_domain} to crawl.")
    else:
        await log_message("❌ No root URL provided.")

    if args.SITEMAP_URL and len(args.SITEMAP_URL) > 0:
        await log_message(f"✅ Sitemap URL provided: {args.SITEMAP_URL}.")
        if not root_domain:
            root_domain = get_root_domain(args.SITEMAP_URL)
        else:
            validate(args.SITEMAP_URL)
        # Fetch URLs from the sitemap
        urls_sitemap = await fetch_sitemap(args.SITEMAP_URL)
        urls.extend(urls_sitemap)
        urls = list(set(urls))  # Remove duplicates
        await log_message(f"✅ Found {len(urls)} pages in the sitemap with root domin {root_domain}.")
    else:
        await log_message("❌ No sitemap URL provided.")

    if len(urls) == 0:
        await log_message("❌ Exiting: No URLs to crawl.")
        return

    await log_message(f"✅ Starting crawl of {len(urls)} urls with max depth {args.MAX_DEPTH}, batch_size={args.BATCH_SIZE}...")
    if args.USER_QUERY is not None:
        await log_message(f"✅ Crawling with user_query={args.USER_QUERY}...")
    force_crawl = True if args.FORCE_CRAWL else False  # Set to True to force crawl of previously crawled pages

    # Start crawling
    await crawl_urls(urls, root_domain, max_depth=args.MAX_DEPTH, batch_size=args.BATCH_SIZE, 
                     user_query=args.USER_QUERY, force_crawl=force_crawl)

    await log_message(f"✅ Crawling complete! Files stored in {OUTPUT_DIR_DEFAULT}")

class UrlValidator(BaseModel):
    url: HttpUrl

def validate(url: str):
    try:
        UrlValidator(url=url)
    except ValidationError as e:
        logger.fatal(f"Exception validating url <{url}>: {e}")
        raise e
    # logger.info(f"validated url: <{url}>")


def get_root_domain(root_url):
    validate(root_url)
    root_url_parsed = urlparse(root_url)
    if not root_url_parsed.scheme or not root_url_parsed.netloc or not root_url_parsed.hostname:
        logger.exception(f'Invalid root_url=<{root_url}>: empty scheme or domain/server or hostname')
        raise ValueError(f'Invalid root_url=<{root_url}>: empty scheme or domain/server or hostname')
    if root_url_parsed.scheme == root_url_parsed.hostname:
        logger.exception(f'Invalid root_url=<{root_url}>: scheme and hostname are same')
        raise ValueError(f'Invalid root_url=<{root_url}>: scheme and hostname are same')
    
    root_domain = root_url_parsed.netloc
    tld_extract_result = tld_extract(root_url)
    root_domain = tld_extract_result.domain + '.' + tld_extract_result.suffix
    logger.info(f'root_domain for url=<{root_url}> = <{root_domain}>')
    return root_domain

def get_args():
    """ Get command line arguments """
    parser = argparse.ArgumentParser()
    # parser.add_argument('-o', '--OUTPUT_DIR', help="output files directory", default=OUTPUT_DIR_DEFAULT)
    parser.add_argument('-su', '--SITEMAP_URL', help="url of sitemap to be crawled", default = SITEMAP_URL_DEFAULT)
    parser.add_argument('-u', '--ROOT_URL', help="root url to be crawled", default = ROOT_URL_DEFAULT)
    # parser.add_argument('-t', '--ROOT_URL_TLD_CRAWL_FLAG', help="flag to indicate to crawl all of the pages matching top level domain of url", required=False)
    parser.add_argument('-d', '--MAX_DEPTH', help="depth of web pages to crawl", default=MAX_DEPTH_DEFAULT, required=False)
    parser.add_argument('-b', '--BATCH_SIZE', help="batch size to crawl web pages in parallel", default=BATCH_SIZE, required=False)
    parser.add_argument('-q', '--USER_QUERY', help="user query to only download web content relevant to the query", required=False)
    # parser.add_argument('-m', '--SAVE_MEDIA_FILES_FLAG', help="flag to save media files (images, audio and video) found on webpages", 
    #                     action='store_true')
    # parser.add_argument('-s', '--SAVE_CRAWL_TO_FILE_FLAG', help="flag to indicate to save list of crawled pages to a file", 
    #                     default=save_crawl_to_file_DEFAULT, required=False)
    parser.add_argument('-f', '--FORCE_CRAWL', help="flag to indicate to force crawl of previously crawled pages", 
                        required=False, default=FORCE_CRAWL, action='store_true')
    # parser.add_argument('-c', '--CONTENT_TYPES', help="regular expression of content types to save when crawling (only matching content types will be saved)", default='', required=False)
    l_args = parser.parse_args()
    # logger.info(vars(l_args))
    # logger.info('\n')
    return l_args
    
# main script run
if __name__ == '__main__':
    logging.basicConfig(level=logging.INFO, format="%(name)s: [%(levelname)s] %(message)s", # format="%(name)s: %(asctime)s: [%(levelname)s] %(message)s",
        handlers=[logging.StreamHandler(), logging.FileHandler(LOGGER_NAME + '.log')])
    logger.info(f'*************** Running python script: {sys.argv} ***************')
    logger.info(f'In working directory: {os.getcwd()}')
    logger.debug(f'Path for loading python modules: {sys.path}')
    
    from dotenv import load_dotenv, find_dotenv
    ENV_FILENAME = '.env'
    SECRETS_FILENAME = '.env.secrets'
    if not load_dotenv(find_dotenv(ENV_FILENAME)):
        logger.warning(f'Failed to load enironment variables from {ENV_FILENAME}')
    else:
        logger.info(f'Loaded enironment variables from {ENV_FILENAME}')
    if not load_dotenv(find_dotenv(SECRETS_FILENAME)):
        logger.warning(f'Failed to load enironment variables from {SECRETS_FILENAME}')
    else:
        logger.info(f'Loaded enironment variables from {SECRETS_FILENAME}')
    
    m_args = get_args()
    #pd.set_option('display.max_rows', 500)
    #pd.set_option('display.max_columns', 500)
    #pd.set_option('display.width', 1000)
    ret = asyncio.run(main(m_args))
    logger.info(f'********** Finished running python script: {sys.argv}: return: {ret} **********')
    sys.exit(0)  # zero exit code to mark Success.
