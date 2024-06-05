#!/usr/bin/env python
"""\
Webpage crawler: given root url, crawl all pages below within same domain and write content of each page to a different file
"""
__author__ = "Ram P"
__email__ = "rambpol@yahoo.com"
__version__ = "0.1.0"
# __maintainer__ = __author__
# __status__ = "Development"
# __copyright__  = "Copyright 2023, " + __author__

import os
import sys
import argparse
import json
import re
import requests
import urllib
import shutil
import time
import string


import logging
LOGGER_NAME =  os.path.splitext(os.path.basename(__file__))[0]
logger = logging.getLogger(name=LOGGER_NAME) # root logger by default, pass LOGGER_NAME for script specific log file
logging.raiseExceptions = False

HTML_EXTENSION = '.html'
JSON_EXTENSION = '.json'
PDF_EXTENSION = '.pdf'
XML_EXTENSION = '.xml'
CONTENT_TYPE_HEADER = 'Content-Type'
PDF_CONTENT_TYPE = 'application/pdf'
HTML_CONTENT_TYPE = 'text/html'
TEXT_CONTENT_TYPE = 'text'
JSON_CONTENT_TYPE = 'application/json'
XML_CONTENT_TYPE = 'application/xml'
IMAGE_CONTENT_TYPES = 'jpeg|jpg'
CRAWL_FILENAME_SUFFIX = '_crawled_pages.xlsx'
VISITED_URLS_COLUMN = 0
VISITED_URL_PATHS_COLUMN = 2
save_crawl_to_file_DEFAULT = True
OUTPUT_DIR_DEFAULT = r'./web_crawler_output'
MAX_DEPTH_DEFAULT = 0
SAVE_MEDIA_FILES_FLAG = False
EXTERNAL_FILE_TYPES_TO_DOWNLOAD = ".pdf | .docx | .xlsx | .pptx | .json "


from bs4 import BeautifulSoup
from urllib.parse import urlparse
from tldextract import extract as tld_extract
from openpyxl import Workbook, load_workbook
# from selenium import webdriver
from requests_html import HTMLSession

from requests.exceptions import ConnectionError, InvalidSchema, ReadTimeout

from pydantic import BaseModel, HttpUrl, ValidationError
class UrlValidator(BaseModel):
    url: HttpUrl

def validate(url: str):
    try:
        UrlValidator(url=url)
    except ValidationError as e:
        logger.fatal(f"Exception validating url <{url}>: {e}")
        raise e
    logger.info(f"Crawling url: <{url}>")

    
def main(args):
    logger.info(f'main(): started with arguments: {args}')
    """ Main logic """
    args.MAX_DEPTH = int(args.MAX_DEPTH)
    args.ROOT_URL_TLD_CRAWL_FLAG = bool(args.ROOT_URL_TLD_CRAWL_FLAG)
    args.SAVE_CRAWL_TO_FILE_FLAG = bool(args.SAVE_CRAWL_TO_FILE_FLAG)
    crawl(args.ROOT_URL, args.ROOT_URL_TLD_CRAWL_FLAG, args.CONTENT_TYPES,
          args.MAX_DEPTH, args.SAVE_MEDIA_FILES_FLAG, args.OUTPUT_DIR, args.SAVE_CRAWL_TO_FILE_FLAG, 
          args.FORCE_CRAWL_FLAG)

def crawl(root_url, crawl_root_url_tld, content_types, max_depth, save_media_files, output_dir, save_crawl_to_file, force_crawl):

    # options = webdriver.ChromeOptions()
    # options.add_argument('--ignore-certificate-errors')
    # options.add_argument('--incognito')
    # options.add_argument('--headless')
    # driver = webdriver.Chrome( executable_path=r'./chromedriver.exe', options=options)
    html_session = HTMLSession()

    root_url = root_url.strip('/')
    validate(root_url)
    root_url_parsed = urlparse(root_url)
    if not root_url_parsed.scheme or not root_url_parsed.netloc or not root_url_parsed.hostname:
        logger.fatal(f'Invalid root_url=<{root_url}>: empty scheme or domain/server or hostname')
        sys.exit(1)
    if root_url_parsed.scheme == root_url_parsed.hostname:
        logger.fatal(f'Invalid root_url=<{root_url}>: scheme and hostname are same')
        sys.exit(1)
    root_domain = root_url_parsed.netloc
    if crawl_root_url_tld:
        # root_tld = '.'.join(root_domain.split('.')[-2:])
        tld_extract_result = tld_extract(root_url)
        root_domain = tld_extract_result.domain + '.' + tld_extract_result.suffix
        logger.info(f'root_url=<{root_url}>, root_domain_tld={root_domain}')
    else:
        logger.info(f'root_url=<{root_url}>, root_domain={root_domain}')
    # to store the pages to be visited
    urls = set(root_domain) 
    url_paths = [root_url] # queue used to get next link to visit
    # to store the pages already visited
    visited_urls = set()
    visited_url_paths = []
    save_dir = os.path.join(output_dir, root_domain.replace('/', '_'))
    if not os.path.exists(save_dir):
        os.makedirs(save_dir)
        saved_file_basenames = set()
    else:
        logger.info(f'Output directory already exists: {save_dir}')
        saved_files = os.listdir(save_dir)
        saved_file_basenames = set([os.path.splitext(saved_file)[0].lower() for saved_file in saved_files])
        logger.info(f'Loaded {len(saved_files)} saved file basenames from directory {save_dir}')

    if save_crawl_to_file:
        crawl_filename = root_domain + '-c_'+ content_types + '-d_'+str(max_depth) + CRAWL_FILENAME_SUFFIX
        crawl_filename = urllib.parse.quote(crawl_filename, safe='', encoding=None, errors=None)
        crawl_filename = os.path.join(output_dir, crawl_filename)

        logger.info(f'crawl_filename={crawl_filename}')
        if os.path.exists(crawl_filename) and not force_crawl:
            wb = load_workbook(crawl_filename)
            ws = wb.active
            total_urls_crawled = ws.max_row - 1
            for row in ws.iter_rows(min_row=2, values_only=True):
                visited_url_paths.append(row[VISITED_URL_PATHS_COLUMN].split(', '))
                visited_urls.add(row[VISITED_URLS_COLUMN])
            logger.info(f'Loaded {total_urls_crawled} previously crawled pages from {crawl_filename}')
            wb.close()
        else:
            wb = Workbook()
            # ws = wb.create_sheet()
            ws = wb.active # default sheet
            ws.append(['URL', 'Content Size', 'URL Path'])
            wb.save(crawl_filename)

    while len(url_paths) > 0:
        current_url_path = url_paths.pop()
        if type(current_url_path) is str:
            current_url_path = [current_url_path]
        current_url = current_url_path[-1]
        save_file_basename = get_save_file_basename(save_dir, current_url)
        save_file_basename_temp = os.path.splitext(os.path.basename(save_file_basename))[0]
        if save_file_basename_temp.lower() in saved_file_basenames and not force_crawl:
            logger.info(f'--Skipping <{current_url}>: already crawled and saved in file with basename: {save_file_basename}')
            continue
        logger.info(f'Visiting link <{current_url}> at depth {len(current_url_path)}: visited: {len(visited_url_paths)}, remaining: {len(url_paths)}')
        try:
            # driver.get(current_url)
            # more_buttons = driver.find_elements_by_class_name("moreLink")
            # for x in range(len(more_buttons)):
            #     if more_buttons[x].is_displayed():
            #         driver.execute_script("arguments[0].click();", more_buttons[x])
            #         time.sleep(1)
            # page_source = driver.page_source
            response = html_session.get(current_url, 
                        # requests.get(current_url, 
                                    headers = {'User-Agent': 'Mozilla/5.0'
                                            #    ,'Client-ID': '<some id>'
                                               }
                                    , verify=False
                                    , timeout=60
                                    # , stream = True   # TODO: add this for urls with media files
                                    )
        except Exception as e:
            logger.error(f'ERROR: exception while fetching {current_url}: {e}')
            continue
        try:
            content_type_header = response.headers.get(CONTENT_TYPE_HEADER)
        except Exception as e:
            logger.error(f'ERROR: exception while parsing header and content in response from {current_url}: {e}')
            continue
        content_type = get_content_type_from_response_header(content_type_header)

        if content_type.lower() in HTML_CONTENT_TYPE or content_type.lower() in JSON_CONTENT_TYPE:
            try:
                response.html.render(sleep = 1, timeout = 10)  # ensures all content in html document is rendered 
                content = response.html.html
            # except InvalidSchema as e:
            #     logger.warning(f'InvalidSchema exception while rendering html content in {current_url}: {e}')
            #     content = response.content
            # except ReadTimeout as e:
            #     logger.warning(f'ReadTimeout exception while rendering html content in {current_url}: {e}')
            #     content = response.content
            # except ConnectionError as e:
            #     logger.warning(f'ConnectionError exception while rendering html content in {current_url}: {e}')
            #     content = response.content
            # except TimeoutError as e:
            #     logger.warning(f'Timeout exception while rendering html content in {current_url}: {e}')
            #     content = response.content
            except Exception as e:
                logger.warning(f'Exception while rendering html content in {current_url}: {e}')
                content = response.content
        else:
            content = response.content

        if len(content_types) > 0 and re.match(content_types, content_type, re.IGNORECASE) is None:
            logger.error(f'--Skipping <{current_url}>: content type {content_type} not matching any of the content types: {content_types}')
        else:
            save_result = save_url_content_to_file(current_url, response, content_type, save_file_basename)
            if save_result and save_crawl_to_file:
                wb = load_workbook(crawl_filename)
                ws = wb.active
                ws.append([current_url, str(len(content)), ', '.join(current_url_path)])
                logger.debug(f'Saved {current_url} to {crawl_filename}')
                wb.save(crawl_filename)
        # mark the current URL as visited
        visited_url_paths.append(current_url_path)
        visited_urls.add(current_url)
        logger.debug(f'--Visited link <{current_url}> at depth {len(current_url_path)}: visited: {len(visited_urls)}, remaining: {len(url_paths)}')
        
        if max_depth > 0 and len(current_url_path) >= max_depth:
            logger.info(f'--Skipping links in <{current_url}>: max depth {max_depth} reached')
            continue

        # if current url is html or json (which is returned by apis), parse and crawl links within
        if not (content_type.lower() in HTML_CONTENT_TYPE or content_type.lower() in JSON_CONTENT_TYPE):
            continue
        try:
            soup = BeautifulSoup(content, "html5lib")
        except Exception as e:
            logger.error(f'Warning: exception parsing content of {current_url} as html: content_type={content_type}: {e}')
            continue
        
        if save_media_files:
            image_srcs = [img['src'] for img in soup.findAll('img')]
            for src in image_srcs:
                r = requests.get(src)
                image_filename = get_save_file_basename(output_dir, src)
                with open(image_filename, 'wb') as file:
                    r.raw.decode_content = True
                    shutil.copyfileobj(r.raw, file)
            audio_links = [a['href'] for a in soup.find_all('a',href=re.compile('http*.*\.(mp3|wav|ogg|wma)'))]
            for audio in audio_links:
                r = requests.get(audio)
                audio_filename = get_save_file_basename(output_dir, src)
                with open(audio_filename, 'wb') as file:
                    r.raw.decode_content = True
                    shutil.copyfileobj(r.raw, file)
            # video_links = soup.find_all('video')['src']
            video_links = re.findall("http*.*.mp4", soup.script.string)
            for video in video_links:
                r = requests.get(video)
                video_filename = get_save_file_basename(output_dir, src)
                with open(video_filename, 'wb') as file:
                    r.raw.decode_content = True
                    shutil.copyfileobj(r.raw, file)
            # patt = re.compile(r'mp4:\s*\["(.+?)"\]')
            # for e in soup.find_all('script'):
            #     m = patt.search(e.string)
            # link_mpeg = soup.select_one('source[type="application/x-mpegURL"]')["src"]
            # link_mp4 = soup.select_one('source[type="video/mp4"]')["src"]
            # TODO: implement media files
            logger.error(f'media file crawling is not implemented yet!!')

        link_elements = soup.select("a[href]")
        for link_element in link_elements:
            url = link_element['href'].strip('\\"/')
            if len(url) == 0 or url == '.':
                continue

            url_parsed = urlparse(url)
            logger.debug(f'-Found link <{url}>: urlparsed: {url_parsed}: in url <{current_url}>')
            if url_parsed.path.endswith(':'):  # things like mailto:
                continue

            if not 'http' in url_parsed.scheme.lower():
                if len(url_parsed.scheme) > 0:
                    continue
                # continue
            if len(url_parsed.netloc) == 0:
                if url.startswith('#'):
                    continue
                else:
                    url = root_url_parsed.scheme + '://' + root_url_parsed.netloc + '/' + url
                    url_parsed = urlparse(url)
                    logger.debug(f'-Reset domain: <{url}>: urlparsed: {url_parsed}: in url <{current_url}>')
            if root_domain in url_parsed.netloc:
                # if the URL discovered is new
                if url not in visited_urls and url not in urls:
                    logger.info(f'---Appending new link <{url}> at depth {len(current_url_path)+1} from url <{current_url}>: queue length={len(url_paths)}')
                    url_path = current_url_path + [url]
                    url_paths.append(url_path)
                    urls.add(url)
                elif url in urls:
                    logger.debug(f'--Discarding link <{url}>: already in queue')
                else:
                    logger.debug(f'--Discarding link <{url}>: already visited')
            else:
                url_path_split = os.path.splitext(url_parsed.path)
                if len(url_path_split[1]) == 0 or not url_path_split[1] in EXTERNAL_FILE_TYPES_TO_DOWNLOAD:
                    logger.debug(f'--Discarding link <{url}>: external domain')
                else: 
                    if url not in visited_urls and url not in urls:
                        logger.info(f'---Appending new link <{url}> at depth {len(current_url_path)+1} from url <{current_url}>: queue length={len(url_paths)}')
                        url_path = current_url_path + [url]
                        url_paths.append(url_path)
                        urls.add(url)
                    else:
                        logger.debug(f'--Discarding link <{url}>: already visited or queued to visit')

def get_save_file_basename(output_dir, current_url):
    url_parsed = urlparse(current_url)
    file_basename = url_parsed.netloc
    url_path_to_file_basename = ''
    if len(url_parsed.path) > 0 and len(url_parsed.path[1:]) > 0:
        if len(url_parsed.query) >0:
            url_path_to_file_basename = urllib.parse.quote(url_parsed.path+'?'+url_parsed.query, safe='', encoding=None, errors=None)
        else:
            url_path_to_file_basename = urllib.parse.quote(url_parsed.path, safe='', encoding=None, errors=None)
        # path_to_file_basename += url_parsed.path.replace('/', '_')
    full_filename = os.path.join(output_dir, file_basename + url_path_to_file_basename)
    return full_filename


def get_content_type_from_response_header(content_type_header):
    if content_type_header is None:
        return 'None'
    content_type = content_type_header.split(';')[0]
    content_type_fields = content_type.split('/')
    if len(content_type_fields) == 1:
        return content_type_fields[0]
    elif content_type_fields[1].lower() == 'application':
        return content_type_fields[0]
    else:
        return content_type_fields[1]

def save_url_content_to_file(current_url, response, content_type, file_basename):
    file_extension = ''
    if (content_type.lower()  in HTML_CONTENT_TYPE) or (len(os.path.splitext(file_basename)[1]) == 0): 
            file_extension = '.' + content_type
    filename = file_basename + file_extension

    if os.path.exists(filename):
        logger.info(f'Already saved content of url <{current_url}> into file <{filename}>')
        return False
    logger.info(f'Writing content of url <{current_url}> into file <{filename}>')
    try:
        if content_type.lower() in PDF_CONTENT_TYPE:
            with open(filename, 'wb') as file:
                for chunk in response.iter_content(chunk_size = 1024):
                    file.write(chunk)
                # response.raw.decode_content = True
                # shutil.copyfileobj(response.raw, file)
        elif content_type.lower() in HTML_CONTENT_TYPE or content_type.lower() in TEXT_CONTENT_TYPE:
            with open(filename, 'w', encoding='utf-8') as file:
                # file.write(str(response.content))
                file.write(str(response.text))
        else: # write as binary
            with open(filename, 'wb') as file:
                for chunk in response.iter_content(chunk_size = 1024):
                    file.write(chunk)
                # response.raw.decode_content = True
                # shutil.copyfileobj(response.raw, file)
    except Exception as e:
        logger.error(f'ERROR: exception writing content of {current_url} into file <{filename}>: {e}')
        return False
    return True


def get_args():
    """ Get command line arguments """
    parser = argparse.ArgumentParser()
    parser.add_argument('-o', '--OUTPUT_DIR', help="output files directory", default=OUTPUT_DIR_DEFAULT)
    parser.add_argument('-u', '--ROOT_URL', help="root url of webpage or website to be crawled", required=True)
    parser.add_argument('-t', '--ROOT_URL_TLD_CRAWL_FLAG', help="flag to indicate to crawl all of the pages matching top level domain of url", required=False)
    parser.add_argument('-d', '--MAX_DEPTH', help="depth of web pages to crawl", default=MAX_DEPTH_DEFAULT, required=False)
    parser.add_argument('-m', '--SAVE_MEDIA_FILES_FLAG', help="flag to save media files (images, audio and video) found on webpages", 
                        action='store_true')
    parser.add_argument('-s', '--SAVE_CRAWL_TO_FILE_FLAG', help="flag to indicate to save list of crawled pages to a file", 
                        default=save_crawl_to_file_DEFAULT, required=False)
    parser.add_argument('-f', '--FORCE_CRAWL_FLAG', help="flag to indicate to force crawl of previously crawled pages", 
                        action='store_true')
    parser.add_argument('-c', '--CONTENT_TYPES', help="regular expression of content types to save when crawling (only matching content types will be saved)", default='', required=False)
    l_args = parser.parse_args()
    # logger.info(vars(l_args))
    # logger.info('\n')
    return l_args
    
# main script run
if __name__ == '__main__':
    logging.basicConfig(level=logging.INFO, format="%(name)s: [%(levelname)s] %(message)s", # format="%(name)s: %(asctime)s: [%(levelname)s] %(message)s",
        handlers=[logging.StreamHandler(), logging.FileHandler(LOGGER_NAME + '.log')])
    logger.info(f'*************** Running python script: {sys.argv} ***************')
    logger.info(f'In working directory: {os.getcwd()}')
    logger.debug(f'Path for loading python modules: {sys.path}')
    
    from dotenv import load_dotenv, find_dotenv
    ENV_FILENAME = '.env'
    SECRETS_FILENAME = '.env.secrets'
    if not load_dotenv(find_dotenv(ENV_FILENAME)):
        logger.warning(f'Failed to load enironment variables from {ENV_FILENAME}')
    else:
        logger.info(f'Loaded enironment variables from {ENV_FILENAME}')
    if not load_dotenv(find_dotenv(SECRETS_FILENAME)):
        logger.warning(f'Failed to load enironment variables from {SECRETS_FILENAME}')
    else:
        logger.info(f'Loaded enironment variables from {SECRETS_FILENAME}')
    
    m_args = get_args()
    #pd.set_option('display.max_rows', 500)
    #pd.set_option('display.max_columns', 500)
    #pd.set_option('display.width', 1000)
    ret = main(m_args)
    logger.info(f'********** Finished running python script: {sys.argv}: return: {ret} **********')
    sys.exit(0)  # zero exit code to mark Success.

